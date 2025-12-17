import datetime
from openpyxl import Workbook
from datetime import datetime
import data.huawei as huawei
import data.oppo as oppo
import data.xiaomi as xiaomi
import data.vivo as vivo


class Comment:
    def __init__(self):
        print("创建表")
        self.wb = Workbook()
        dt = datetime.now().strftime('%Y-%m-%d')
        self.filename = f"{dt}评论.xlsx"
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

    # 小米评论
    def XiaoMiComment(self):
        data = xiaomi.getContent()
        rows = []
        comment = ["时间", "用户", "内容", "评论星", "机型", "版本", "版本号"]
        rows.append(comment)
        if data and data["data"]["commentInfoList"]:
            for comment in data["data"]["commentInfoList"]:
                record = comment['main']
                print(
                    f"时间:{record['updateTime']} 用户:{record['userName']} 内容:{record['content']} 评论星:{record['score']} 机型:{record['device']} 版本:{record['versionName']} 版本号:{record['versionCode']} \n")
                re = [self.handlerDate(record['updateTime']), record['userName'], record['content'], record['score'],
                      record['device'],
                      record['versionName'], record['versionCode']]
                rows.append(re)
        self.saveDataToExcel(rows, "小米评论", 1)

    # 小米评分
    def XiaoMiScore(self):
        score = xiaomi.getCore()["data"]["score"]
        print(f"评分：{score['score']} 评论数：{score['count']} 5星：{score['count5']} 4星：{score['count4']} 3星：{score['count3']} 2星：{score['count2']} 1星：{score['count1']}")
        row = ["小米", score['score'], score['count'], score['count5'], score['count4'], score['count3'],score['count2'],score['count1']]
        return row

    # vivo 评论
    def ViVoComment(self):
        record = vivo.getContent()
        rows = []
        comment = ["时间", "用户", "评论星", "内容", "机型", "APP版本", "点赞数"]
        rows.append(comment)
        if record and record["data"]["data"]:
            for row in record["data"]["data"]:
                print(f"时间：{row['commentOn']} 用户：{row['userName']} 星级：{row['score']} 内容：{row['comment']} 机型：{row['model']} APP版本：{row['appVersion']} 点赞数：{row['goodCount']} \n")
                d = [self.handlerDate(row['commentOn']), row['userName'], row['score'], row['comment'], row['model'],row['appVersion'], row['goodCount']]
                rows.append(d)
        self.saveDataToExcel(rows, "VIVO评论", 2)

    # vivo 评分
    def viVoScore(self):
        data = vivo.getCore()
        data = data["data"]
        if data:
            print(f"星级：{data['starRating']}, 评分个数：{data['socreCommentCount']} 评论个数：{data['commentCount']}")
            return ["VIVO", round(data['starRating'],1), data['commentCount']]

    # oppo 评论
    def OppoComment(self):
        data = oppo.getContent()
        rows = []
        comment = ["时间", "用户", "评分", "内容", "机型"]
        rows.append(comment)
        if data and data["data"]["rows"]:
            for row in data["data"]["rows"]:
                print(f"时间：{row['UPDATE_TIME']} 用户：{row['USER_NICKNAME']} 评分：{row['USER_GRADE']} 内容：{row['PM_WORD']} 机型：{row['MOBILE_NAME']} \n")
                d = [row['UPDATE_TIME'],row['USER_NICKNAME'],row['USER_GRADE'],row['PM_WORD'],row['MOBILE_NAME']]
                rows.append(d)
        self.saveDataToExcel(rows, "OPPO评论", 3)

    def OppoScore(self):
        data = oppo.getCore()["data"]
        detail = data["score_detail"]
        print(f"评分：{data['agv_score']} 评论数：{data['count']} 5星：{detail['5']['count']} 4星：{detail['4']['count']} 3星：{detail['3']['count']} 2星：{detail['2']['count']} 1星：{detail['1']['count']}")
        row = ["OPPO",round(data['agv_score'],1),data['count'],detail['5']['count'],detail['4']['count'],detail['3']['count'],detail['2']['count'],detail['1']['count']]
        return row
    # 华为评论
    def HuaweiComment(self):
        data = huawei.getContent()
        if data and data["data"]["reviewList"]:
            data =data['data']['reviewList']
            rows = []
            comment = ["时间", "用户", "评分", "内容", "机型","版本"]
            rows.append(comment)
            for row in data:
                d = [self.handlerDate(row['operTimeStamp']), row['nickName'], row['rating'], row['content'], row['phoneType'],row["version"]]
                rows.append(d)
            self.saveDataToExcel(rows, "huawei评论", 4)

    # 华为评分
    def HuaweiScore(self):
        data = huawei.getCore()
        data =data['data']
        print(f"评分：{data['star']} 评论数：{data['totalReviews']}")
        row = ["华为",round(data['star'],1),data['totalReviews'],None,None,None,None,None]
        return row

    # 插入数据
    def saveDataToExcel(self, data, tableName, index):
        ws = self.wb.create_sheet(tableName, index)
        for k in data:
            ws.append(k)
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母（如 'A', 'B'）
            for cell in col:
                try:
                    # 计算单元格内容的字符长度
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            # 设置列宽（增加 2 个字符的缓冲）
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def handlerDate(self, timestamp):
        timestamp_seconds = timestamp / 1000
        dt = datetime.fromtimestamp(timestamp_seconds)
        formatted_date = dt.strftime('%Y-%m-%d')
        return formatted_date

    def saveScoreToExcel(self):
        rows = []
        rows.append(["平台", "评分", "评论数", "5星", "4星", "3星", "2星", "1星"])
        rows.append(self.XiaoMiScore())
        rows.append(self.viVoScore())
        rows.append(self.OppoScore())
        rows.append(self.HuaweiScore())
        self.saveDataToExcel(rows, "总评分", 0)

    def getDataFromExcel(self):
        self.XiaoMiComment()
        self.ViVoComment()
        self.OppoComment()
        self.HuaweiComment()
        self.saveScoreToExcel()
        self.wb.save(self.filename)


if __name__ == '__main__':
    client = Comment()
    client.getDataFromExcel()
